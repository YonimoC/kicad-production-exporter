"""
Production File Exporter - KiCad 生产文件一键导出插件

SWIG Action Plugin，兼容 KiCad 8.0 ~ 10.0+（自动检测版本适配）。
在 PCB 编辑器中一键导出所有生产文件。

安装：将此文件及 icon.png 放入 KiCad 插件目录：
  Windows: %USERPROFILE%\Documents\KiCad\<版本>\scripts\plugins\
  macOS:   ~/Documents/KiCad/<版本>/scripts/plugins/
  Linux:   ~/.local/share/kicad/<版本>/scripts/plugins/
"""

import os
import re
import sys
import subprocess
from pathlib import Path

import wx
import pcbnew

# ---- 模块级配置 ----
# Windows 下隐藏命令行窗口（兼容旧版 Python）
_CREATE_FLAGS = getattr(subprocess, 'CREATE_NO_WINDOW', 0) if sys.platform == "win32" else 0

# ---- 模块级版本缓存 ----
_kicad_version_cache: tuple | None = None


def find_kicad_cli(kicad_ver: tuple | None = None) -> str:
    """查找 kicad-cli，优先用当前 KiCad 自带的。"""
    import shutil

    # 方式 1: 从当前运行的 KiCad 的 Python 路径推断（最可靠）
    # sys.executable 是 KiCad 自带的 python.exe，在 .../KiCad/10.0/bin/python.exe
    kicad_bin = os.path.dirname(sys.executable)
    cli_exe = os.path.join(kicad_bin, "kicad-cli.exe" if sys.platform == "win32" else "kicad-cli")
    if os.path.exists(cli_exe):
        print(f"[ProductionExporter] kicad-cli: {cli_exe} (当前KiCad)")
        return cli_exe

    # 方式 2: PATH
    path_cli = shutil.which("kicad-cli")
    if path_cli:
        return path_cli

    # 方式 3: 指定版本目录
    if kicad_ver and sys.platform == "win32":
        p = rf"C:\Program Files\KiCad\{kicad_ver[0]}.0\bin\kicad-cli.exe"
        if os.path.exists(p):
            return p

    # 方式 4: 常见目录（新版优先）
    if sys.platform == "win32":
        for ver in ["11.0", "10.0", "9.0", "8.0"]:
            p = rf"C:\Program Files\KiCad\{ver}\bin\kicad-cli.exe"
            if os.path.exists(p):
                return p

    return "kicad-cli"


def detect_kicad_version() -> tuple[int, int]:
    """
    检测当前 KiCad 版本，返回 (major, minor) 元组，如 (8, 0) 或 (10, 0)。
    结果会被缓存，多次调用只执行一次检测。
    """
    global _kicad_version_cache
    if _kicad_version_cache is not None:
        return _kicad_version_cache

    # 方式 1: 当前运行的 pcbnew 模块
    for method in ["GetBuildVersion", "Version", "GetVersion"]:
        try:
            ver_str = str(getattr(pcbnew, method)())
            print(f"[ProductionExporter] pcbnew.{method}() = {ver_str}")
            m = re.search(r"(\d+)\.(\d+)", ver_str)
            if m:
                _kicad_version_cache = (int(m.group(1)), int(m.group(2)))
                return _kicad_version_cache
        except Exception as e:
            print(f"[ProductionExporter] pcbnew.{method}() 失败: {e}")
            continue

    # 方式 2: kicad-cli --version（可能找到其他版本）
    try:
        result = subprocess.run(
            [find_kicad_cli(), "--version"],
            capture_output=True, text=True, timeout=10,
            creationflags=_CREATE_FLAGS
        )
        if result.returncode == 0:
            m = re.search(r"(\d+)\.(\d+)", result.stdout)
            if m:
                _kicad_version_cache = (int(m.group(1)), int(m.group(2)))
                return _kicad_version_cache
    except Exception:
        pass

    # 无法检测
    print("[ProductionExporter] ⚠ 无法检测 KiCad 版本，假定为 8.0")
    _kicad_version_cache = (8, 0)
    return _kicad_version_cache


class ProductionExporter(pcbnew.ActionPlugin):
    """一键导出生产文件插件"""

    def defaults(self):
        self.name = "导出生产文件"
        self.category = "生产制造"
        self.description = (
            "一键导出 GERBER、丝印图、3D视图PNG、坐标文件与 BOM"
        )
        self.show_toolbar_button = True
        self.icon_file_name = os.path.join(
            os.path.dirname(__file__), "icon.png"
        )

    def Run(self):
        """插件入口：用户点击工具栏按钮时调用。"""
        # ---- 0. 检测 KiCad 版本 ----
        ver = detect_kicad_version()
        print(f"[ProductionExporter] 检测到 KiCad {ver[0]}.{ver[1]}")

        # ---- 1. 获取当前 PCB 文件路径 ----
        board = pcbnew.GetBoard()
        board_path = board.GetFileName()

        if not board_path:
            self._show_error("错误", "请先打开一个 PCB 文件再使用此插件。")
            return

        project_name = Path(board_path).stem

        # ---- 2. 选择要导出的项目（默认全选）----
        export_items = self._select_export_items(ver)
        if not export_items:
            return  # 用户取消了

        # ---- 3. 选择输出文件夹 ----
        output_root = self._select_output_folder()
        if not output_root:
            return  # 用户取消了

        # ---- 4. 确认导出 ----
        # 提前获取板层信息用于显示
        copper_count, copper_names = self._get_copper_info()
        # 格式化铜层列表（每行不超过 ~80 字符，多了就换行）
        cu_str = ", ".join(copper_names)
        all_layers = copper_names + [
            "F.Paste", "B.Paste", "F.Silkscreen", "B.Silkscreen",
            "F.Mask", "B.Mask", "Edge.Cuts"
        ]
        layer_summary = (
            f"识别为 {copper_count}层板\n"
            f"  铜层({copper_count}): {cu_str}\n"
            f"  GERBER 共导出 {len(all_layers)} 层"
        )

        mode = f"KiCad {ver[0]}.{ver[1]}" + (" (K8模式)" if ver[0] < 9 else " (K9+模式)")
        # 统计选中项目
        selected_names = [n for n, sel in export_items.items() if sel]
        items_str = "\n".join(f"  ├── {n}" for n in selected_names[:-1]) if len(selected_names) > 1 else ""
        if len(selected_names) > 1:
            items_str += f"\n  └── {selected_names[-1]}"
        elif selected_names:
            items_str = f"  └── {selected_names[0]}"

        dlg = wx.MessageDialog(
            None,
            f"将导出以下内容到:\n{output_root}\n\n"
            f"{layer_summary}\n\n"
            f"检测版本: {mode}\n"
            f"选中项目 ({len(selected_names)}):\n"
            f"{items_str}\n\n"
            f"是否继续？",
            "确认导出",
            wx.YES_NO | wx.ICON_QUESTION
        )
        if dlg.ShowModal() != wx.ID_YES:
            dlg.Destroy()
            return
        dlg.Destroy()

        # ---- 5. 执行导出（带进度条）----
        selected_count = sum(1 for sel in export_items.values() if sel)
        total = selected_count
        if total == 0:
            return
        dlg = wx.ProgressDialog(
            "导出生产文件", "准备中...", maximum=total,
            style=wx.PD_AUTO_HIDE | wx.PD_APP_MODAL | wx.PD_ELAPSED_TIME | wx.PD_REMAINING_TIME)
        dlg.SetSize(450, 180)

        def on_progress(step: int, desc: str):
            dlg.Update(step, f"[{step}/{total}] {desc}")

        try:
            self._do_export(board_path, output_root, project_name, ver, export_items, on_progress, total)
        finally:
            dlg.Destroy()

    # ================================================================
    #  内部方法
    # ================================================================

    def _select_export_items(self, kicad_ver: tuple) -> dict | None:
        """弹出导出项目选择对话框，返回 {项目名: 是否选中} 或 None（取消）。"""
        # 定义所有可选导出项目
        all_items = [
            ("GERBER 文件", True),
            ("钻孔文件 (PTH+NPTH)", True),
            ("顶层丝印图 (PDF)", True),
            ("底层丝印图 (PDF)", True),
            ("坐标文件 (CSV)", True),
            ("BOM 物料清单 (XLSX)", True),
        ]
        if kicad_ver[0] >= 9:
            all_items += [
                ("3D 顶层视图 (PNG)", True),
                ("3D 底层视图 (PNG)", True),
            ]
        all_items.append(("3D STEP 模型", True))

        # 构建自定义对话框
        dlg = wx.Dialog(None, title="选择导出项目",
                        size=(380, 420),
                        style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER)
        panel = wx.Panel(dlg)
        vbox = wx.BoxSizer(wx.VERTICAL)

        title = wx.StaticText(panel, label="请选择要导出的项目：")
        title_font = title.GetFont()
        title_font.SetWeight(wx.FONTWEIGHT_BOLD)
        title.SetFont(title_font)
        vbox.Add(title, 0, wx.ALL, 10)

        checkboxes = []
        for name, default in all_items:
            cb = wx.CheckBox(panel, label=name)
            cb.SetValue(default)
            checkboxes.append(cb)
            vbox.Add(cb, 0, wx.LEFT | wx.RIGHT | wx.TOP, 20)

        # 全选/全不选按钮
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        btn_all = wx.Button(panel, label="全选")
        btn_none = wx.Button(panel, label="全不选")
        btn_sizer.Add(btn_all, 0, wx.RIGHT, 10)
        btn_sizer.Add(btn_none, 0)
        vbox.Add(btn_sizer, 0, wx.LEFT | wx.RIGHT | wx.TOP, 20)

        def on_select_all(e):
            for cb in checkboxes:
                cb.SetValue(True)

        def on_select_none(e):
            for cb in checkboxes:
                cb.SetValue(False)

        btn_all.Bind(wx.EVT_BUTTON, on_select_all)
        btn_none.Bind(wx.EVT_BUTTON, on_select_none)

        # 确定/取消
        btn_ok = wx.Button(panel, wx.ID_OK, "确定")
        btn_cancel = wx.Button(panel, wx.ID_CANCEL, "取消")
        ok_sizer = wx.BoxSizer(wx.HORIZONTAL)
        ok_sizer.AddStretchSpacer()
        ok_sizer.Add(btn_ok, 0, wx.RIGHT, 10)
        ok_sizer.Add(btn_cancel, 0)
        vbox.Add(ok_sizer, 0, wx.ALL | wx.EXPAND, 15)

        panel.SetSizer(vbox)
        dlg.Fit()
        dlg.CenterOnScreen()

        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return None
        dlg.Destroy()

        # 构建返回结果
        result = {}
        for (name, _), cb in zip(all_items, checkboxes):
            result[name] = cb.GetValue()
        return result

    def _select_output_folder(self) -> str | None:
        """弹出文件夹选择对话框，返回所选路径或 None。"""
        # wx.DD_DIR_MUST_EXIST 可能不存在于旧版 wx
        style = wx.DD_DEFAULT_STYLE
        if hasattr(wx, 'DD_DIR_MUST_EXIST'):
            style |= wx.DD_DIR_MUST_EXIST
        dlg = wx.DirDialog(None, "选择生产文件输出文件夹", style=style)
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            dlg.Destroy()
            return path
        dlg.Destroy()
        return None

    def _show_info(self, title: str, message: str):
        """信息提示框。"""
        wx.MessageBox(message, title, wx.OK | wx.ICON_INFORMATION)

    def _show_error(self, title: str, message: str):
        """错误提示框。"""
        wx.MessageBox(message, title, wx.OK | wx.ICON_ERROR)

    def _show_warning(self, title: str, message: str):
        """警告提示框。"""
        wx.MessageBox(message, title, wx.OK | wx.ICON_WARNING)

    def _get_logger(self, output_root: str):
        """返回一个日志函数，写入输出目录下的 _export.log。"""
        log_path = os.path.join(output_root, "_export.log")
        def log(msg: str):
            try:
                timestamp = __import__('datetime').datetime.now().strftime("%H:%M:%S")
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(f"[{timestamp}] {msg}\n")
            except Exception:
                pass
        return log

    def _try_cmd(self, cmd: list, description: str) -> tuple[bool, str]:
        """执行单个命令。kicad-cli 输出的"Warning/警告"不作为错误。"""
        print(f"[ProductionExporter] 命令: {' '.join(cmd)}")
        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True, timeout=300,
                creationflags=_CREATE_FLAGS
            )
            stderr = (result.stderr or "").strip()
            stdout = (result.stdout or "").strip()

            if result.returncode == 0:
                if stderr:
                    print(f"[ProductionExporter] ✅ {description} 完成 (有警告)")
                else:
                    print(f"[ProductionExporter] ✅ {description} 完成")
                return True, ""

            # 非零退出码，检查是否只有警告（没有真正的错误）
            combined = (stderr + stdout).strip()
            # 过滤掉已知的非致命警告行
            warn_lines = []
            real_errors = []
            for line in combined.splitlines():
                line_stripped = line.strip()
                if not line_stripped:
                    continue
                # 跳过时间戳 + Warning/警告 的行
                if re.match(r'^\d+:\d+:\d+:\s*(Warning|警告)', line_stripped):
                    warn_lines.append(line_stripped)
                    continue
                # 跳过纯 library table 警告
                if 'Library table' in line_stripped and ('UNINITIALIZED' in line_stripped or 'skipping' in line_stripped):
                    warn_lines.append(line_stripped)
                    continue
                real_errors.append(line_stripped)

            if not real_errors and warn_lines:
                # 只有警告，没有真正的错误 → 视作成功
                print(f"[ProductionExporter] ✅ {description} 完成 (忽略 {len(warn_lines)} 条警告)")
                return True, ""

            err = "\n".join(real_errors) or f"返回码 {result.returncode}"
            short_err = err[:200]
            print(f"[ProductionExporter] ❌ {description}: {short_err}")
            return False, short_err
        except FileNotFoundError:
            return False, "找不到 kicad-cli"
        except subprocess.TimeoutExpired:
            return False, f"{description} 超时"
        except Exception as e:
            return False, str(e)[:200]

    def _run_with_fallback(self, primary: list, fallback: list,
                           description: str) -> tuple[bool, str]:
        """先尝试主命令，失败则尝试备选命令。"""
        ok, err = self._try_cmd(primary, description)
        if ok:
            return True, ""
        print(f"[ProductionExporter] 主命令失败，尝试备选...")
        ok2, err2 = self._try_cmd(fallback, description + " (备选)")
        return ok2, err2

    def _try_drill_export(self, cli: str, gdir: str, board_path: str) -> tuple[bool, str]:
        """钻孔导出：优先分离 PTH/NPTH，不支持则兜底合并。"""
        # --excellon-separate-th 是 KiCad 官方标志（从 8.0 起支持）
        ok, err = self._try_cmd(
            [cli, "pcb", "export", "drill",
             "--excellon-separate-th", "-o", gdir, board_path],
            "钻孔 (分离PTH/NPTH)")
        if ok:
            return True, ""
        # 不支持分离则兜底
        ok2, err2 = self._try_cmd(
            [cli, "pcb", "export", "drill", "-o", gdir, board_path],
            "钻孔 (不分离)")
        return ok2, err2

    def _get_copper_info(self) -> tuple[int, list]:
        """检测当前 PCB 的铜层信息，返回 (铜层数, [铜层名列表])。"""
        try:
            board = pcbnew.GetBoard()
            count = board.GetCopperLayerCount()
            copper = []
            for i in range(count):
                # 铜层 ID: F_Cu(0) ... In(N-2)_Cu(N-2), B_Cu(31)
                # 内层连续递增，底层是固定的 B_Cu
                lid = pcbnew.B_Cu if i == count - 1 else pcbnew.F_Cu + i
                copper.append(board.GetLayerName(lid))
            return len(copper), copper
        except Exception as e:
            print(f"[ProductionExporter] 铜层检测失败: {e}")
            return 2, ["F.Cu", "B.Cu"]

    def _get_gerber_layers(self) -> str:
        """动态获取 GERBER 导出所需的全部图层（含所有铜层）。
        使用 kicad-cli 识别的规范层名，而非 board.GetLayerName() 可能返回的用户自定义名。"""
        _, copper_names = self._get_copper_info()
        # 强制使用 kicad-cli 标准层名（规避用户重命名导致 kicad-cli 不识别）
        canonical_copper = []
        count = len(copper_names)
        for i in range(count):
            if i == 0:
                canonical_copper.append("F.Cu")
            elif i == count - 1:
                canonical_copper.append("B.Cu")
            else:
                canonical_copper.append(f"In{i}.Cu")
        layers = list(canonical_copper)
        layers.extend([
            "F.Paste", "B.Paste",
            "F.Silkscreen", "B.Silkscreen",
            "F.Mask", "B.Mask",
            "Edge.Cuts",
        ])
        return ",".join(layers)

    # ================================================================
    #  版本分离的导出方法
    # ================================================================

    def _do_export(self, board_path: str, output_root: str, project_name: str,
                   kicad_ver: tuple, export_items: dict, on_progress, total: int):
        """根据 KiCad 版本选择对应的导出方法。"""
        if kicad_ver[0] < 9:
            self._do_export_v8(board_path, output_root, project_name, export_items, on_progress, total)
        else:
            self._do_export_v9(board_path, output_root, project_name, kicad_ver, export_items, on_progress, total)

    # ----------------------------------------------------------------
    #  KiCad 8 导出
    # ----------------------------------------------------------------
    def _do_export_v8(self, board_path: str, output_root: str,
                      project_name: str, export_items: dict, on_progress, total: int):
        """KiCad 8.x 导出命令（短标志风格）。"""
        cli = find_kicad_cli()
        bf = Path(board_path)
        results: list[tuple[str, bool, str]] = []
        sch = bf.with_suffix(".kicad_sch")

        gdir = os.path.join(output_root, f"{project_name}_Gerber")
        sdir = os.path.join(output_root, f"{project_name}_丝印与3D图")
        cdir = os.path.join(output_root, f"{project_name}_坐标文件")

        for d in (gdir, sdir, cdir):
            os.makedirs(d, exist_ok=True)

        print(f"[ProductionExporter] K8 模式 | 项目: {project_name}")

        def run(cmd, desc):
            try:
                return self._try_cmd(cmd, desc)
            except Exception as e:
                return False, str(e)[:200]

        step = 0

        # -- GERBER --
        if export_items.get("GERBER 文件", True):
            step += 1; on_progress(step, "正在导出 GERBER 文件...")
            layers = self._get_gerber_layers()
            ok, err = self._try_cmd(
                [cli, "pcb", "export", "gerbers",
                 "-l", layers, "--output", gdir, str(bf)],
                "GERBER 文件导出")
            if not ok:
                ok2, err2 = self._try_cmd(
                    [cli, "pcb", "export", "gerbers",
                     str(bf), "--output", gdir, "-l", layers],
                    "GERBER (板文件在前)")
                if ok2: ok, err = True, ""
                else:
                    ok3, err3 = self._try_cmd(
                        [cli, "pcb", "export", "gerbers",
                         "--output", gdir, str(bf)],
                        "GERBER (无层)")
                    if ok3: ok, err = True, ""
            results.append(("GERBER 文件", ok, err))

        # -- 钻孔 --
        if export_items.get("钻孔文件 (PTH+NPTH)", True):
            step += 1; on_progress(step, "正在导出钻孔文件...")
            ok, err = self._try_drill_export(cli, gdir, str(bf))
            results.append(("钻孔文件 (PTH+NPTH)", ok, err))

        # -- 丝印 PDF --
        for side, cn in [("F", "顶层"), ("B", "底层")]:
            key = f"{cn}丝印图 (PDF)"
            if export_items.get(key, True):
                step += 1; on_progress(step, f"正在导出{cn}丝印图...")
                out = os.path.join(sdir, f"{project_name}_{cn}丝印.pdf")
                is_bot = (side == "B")
                cmd_enhanced = [cli, "pcb", "export", "pdf",
                                "-l", f"{side}.Paste,{side}.Silkscreen,{side}.Mask,Edge.Cuts",
                                "--black-and-white",
                                *(["--mirror"] if is_bot else []),
                                "-o", out, str(bf)]
                cmd_base = [cli, "pcb", "export", "pdf",
                            "-l", f"{side}.Paste,{side}.Silkscreen,{side}.Mask,Edge.Cuts",
                            "-o", out, str(bf)]
                ok, err = self._run_with_fallback(cmd_enhanced, cmd_base, f"{cn}丝印图")
                results.append((key, ok, err))

        # -- 坐标 --
        if export_items.get("坐标文件 (CSV)", True):
            step += 1; on_progress(step, "正在导出坐标文件...")
            ok, err = run([cli, "pcb", "export", "pos",
                           "-o", os.path.join(cdir, f"{project_name}_坐标.csv"),
                           "--format", "csv", "--units", "mm", str(bf)],
                          "坐标文件")
            results.append(("坐标文件 (CSV)", ok, err))

        # -- BOM --
        if export_items.get("BOM 物料清单 (XLSX)", True):
            step += 1; on_progress(step, "正在导出 BOM...")
            if sch.exists():
                self._bom_export(cli, sch, output_root, project_name, True, results)
            else:
                results.append(("BOM 物料清单", False, "未找到原理图"))

        # -- 3D 视图 (K8 不支持) --
        for key in ["3D 顶层视图 (PNG)", "3D 底层视图 (PNG)"]:
            if export_items.get(key, True):
                results.append((key, False, "需 KiCad 9+"))

        # -- 3D STEP --
        if export_items.get("3D STEP 模型", True):
            step += 1; on_progress(step, "正在导出 3D STEP 模型...")
            ok, err = run([cli, "pcb", "export", "step",
                           "-o", os.path.join(sdir, f"{project_name}_3D.step"),
                           str(bf)], "3D STEP 模型")
            results.append(("3D STEP 模型", ok, err))

        self._print_summary(results, output_root)

    # ----------------------------------------------------------------
    #  KiCad 9+ 导出
    # ----------------------------------------------------------------
    def _do_export_v9(self, board_path: str, output_root: str,
                      project_name: str, kicad_ver: tuple, export_items: dict,
                      on_progress, total: int):
        """KiCad 9.x+ 导出命令（长标志风格，3D 渲染）。"""
        cli = find_kicad_cli(kicad_ver)
        bf = Path(board_path)
        results: list[tuple[str, bool, str]] = []
        sch = bf.with_suffix(".kicad_sch")

        gdir = os.path.join(output_root, f"{project_name}_Gerber")
        sdir = os.path.join(output_root, f"{project_name}_丝印与3D图")
        cdir = os.path.join(output_root, f"{project_name}_坐标文件")

        for d in (gdir, sdir, cdir):
            os.makedirs(d, exist_ok=True)

        print(f"[ProductionExporter] K9+ 模式 | 项目: {project_name}")

        def run(cmd, desc):
            try:
                return self._try_cmd(cmd, desc)
            except Exception as e:
                return False, str(e)[:200]

        step = 0

        # -- GERBER --
        if export_items.get("GERBER 文件", True):
            step += 1; on_progress(step, "正在导出 GERBER 文件...")
            layers = self._get_gerber_layers()
            ok, err = self._try_cmd(
                [cli, "pcb", "export", "gerbers",
                 "-l", layers, "--output", gdir, str(bf)],
                "GERBER 文件导出")
            if not ok:
                ok2, err2 = self._try_cmd(
                    [cli, "pcb", "export", "gerbers",
                     str(bf), "--output", gdir, "-l", layers],
                    "GERBER (板文件在前)")
                if ok2: ok, err = True, ""
                else:
                    ok3, err3 = self._try_cmd(
                        [cli, "pcb", "export", "gerbers",
                         "--output", gdir, str(bf)],
                        "GERBER (无层)")
                    if ok3: ok, err = True, ""
            results.append(("GERBER 文件", ok, err))

        # -- 钻孔 --
        if export_items.get("钻孔文件 (PTH+NPTH)", True):
            step += 1; on_progress(step, "正在导出钻孔文件...")
            ok, err = self._try_drill_export(cli, gdir, str(bf))
            results.append(("钻孔文件 (PTH+NPTH)", ok, err))

        # -- 丝印 PDF --
        for side, cn in [("F", "顶层"), ("B", "底层")]:
            key = f"{cn}丝印图 (PDF)"
            if export_items.get(key, True):
                step += 1; on_progress(step, f"正在导出{cn}丝印图...")
                out = os.path.join(sdir, f"{project_name}_{cn}丝印.pdf")
                is_bot = (side == "B")
                cmd_enhanced = [cli, "pcb", "export", "pdf",
                                "-l", f"{side}.Paste,{side}.Silkscreen,{side}.Mask,Edge.Cuts",
                                "--black-and-white",
                                *(["--mirror"] if is_bot else []),
                                "-o", out, str(bf)]
                cmd_base = [cli, "pcb", "export", "pdf",
                            "-l", f"{side}.Paste,{side}.Silkscreen,{side}.Mask,Edge.Cuts",
                            "-o", out, str(bf)]
                ok, err = self._run_with_fallback(cmd_enhanced, cmd_base, f"{cn}丝印图")
                results.append((key, ok, err))

        # -- 坐标 --
        if export_items.get("坐标文件 (CSV)", True):
            step += 1; on_progress(step, "正在导出坐标文件...")
            ok, err = run([cli, "pcb", "export", "pos",
                           "--output", os.path.join(cdir, f"{project_name}_坐标.csv"),
                           "--format", "csv", "--units", "mm", str(bf)],
                          "坐标文件")
            results.append(("坐标文件 (CSV)", ok, err))

        # -- BOM --
        if export_items.get("BOM 物料清单 (XLSX)", True):
            step += 1; on_progress(step, "正在导出 BOM...")
            if sch.exists():
                self._bom_export(cli, sch, output_root, project_name, False, results)
            else:
                results.append(("BOM 物料清单", False, "未找到原理图"))

        # -- 3D 视图 --
        for side, cn in [("top", "顶层"), ("bottom", "底层")]:
            key = f"3D {cn}视图 (PNG)"
            if export_items.get(key, True):
                step += 1; on_progress(step, f"正在导出3D{cn}视图...")
                out = os.path.join(sdir, f"{project_name}_3D_{cn}.png")
                ok, err = self._run_with_fallback(
                    [cli, "pcb", "render",
                     "--side", side, "--output", out, str(bf)],
                    [cli, "pcb", "export", "pdf",
                     "-l", (f"F.Cu,F.Mask,F.Paste,F.Silkscreen,Edge.Cuts" if side == "top"
                            else f"B.Cu,B.Mask,B.Paste,B.Silkscreen,Edge.Cuts"),
                     "-o", out, str(bf)],
                    f"3D {cn}视图")
                results.append((key, ok, err))

        # -- 3D STEP --
        if export_items.get("3D STEP 模型", True):
            step += 1; on_progress(step, "正在导出 3D STEP 模型...")
            ok, err = run([cli, "pcb", "export", "step",
                           "-o", os.path.join(sdir, f"{project_name}_3D.step"),
                           str(bf)], "3D STEP 模型")
            results.append(("3D STEP 模型", ok, err))

        self._print_summary(results, output_root)

    # ================================================================
    #  工具方法
    # ================================================================

    def _bom_export(self, cli: str, sch, output_root: str, project_name: str,
                    is_v8: bool, results: list):
        """导出 BOM CSV 并转为格式化的 XLSX。"""
        bom_csv = os.path.join(output_root, f"bom-{project_name}.csv")
        flag = "-o" if is_v8 else "--output"
        log = self._get_logger(output_root)

        # QUANTITY/DNP 是 KiCad 内置字段（K8+ 均支持），配合 --group-by 自动按分组计数
        # 注意：不在 CLI 中排序，由 Python 读取 CSV 后按 DNP 排序
        ok, err = self._try_cmd(
            [cli, "sch", "export", "bom",
             "--fields", "Reference,Value,Footprint,QUANTITY,Description,PartNumber,DNP",
             "--labels", "Refs,Value,Footprint,Qty,Description,PartNumber,DNP",
             "--group-by", "Value,Footprint,DNP",
             flag, bom_csv, str(sch)],
            "BOM 导出")
        log(f"BOM 主命令: ok={ok}, err={err}, csv_exists={os.path.exists(bom_csv)}")
        if not ok or not os.path.exists(bom_csv):
            # 备选1: 不带 QUANTITY/DNP，靠 _count_refs 手动计算
            ok2, err2 = self._try_cmd(
                [cli, "sch", "export", "bom",
                 "--fields", "Reference,Value,Footprint,Description,PartNumber",
                 "--labels", "Refs,Value,Footprint,Description,PartNumber",
                 "--group-by", "Value,Footprint",
                 flag, bom_csv, str(sch)],
                "BOM 导出(备选1)")
            log(f"BOM 备选1: ok={ok2}, err={err2}, csv_exists={os.path.exists(bom_csv)}")
            if ok2 and os.path.exists(bom_csv):
                ok, err = True, ""
        if not ok or not os.path.exists(bom_csv):
            # 备选2: 默认参数兜底
            ok3, err3 = self._try_cmd(
                [cli, "sch", "export", "bom", flag, bom_csv, str(sch)],
                "BOM 导出(备选2)")
            log(f"BOM 备选2: ok={ok3}, err={err3}, csv_exists={os.path.exists(bom_csv)}")
            if ok3 and os.path.exists(bom_csv):
                ok, err = True, ""

        if ok and os.path.exists(bom_csv):
            xlsx = os.path.join(output_root, f"bom-{project_name}.xlsx")
            gen_ok, gen_msg = self._generate_bom_xlsx(bom_csv, xlsx)
            if gen_ok:
                try:
                    os.remove(bom_csv)
                except OSError:
                    pass
            else:
                err = gen_msg
                # 保留 CSV 作为兜底
                results.append(("BOM 物料清单 (CSV)", ok, err))
                return
        results.append(("BOM 物料清单 (XLSX)", ok, err))

    def _print_summary(self, results: list, output_root: str):
        """打印并弹窗显示最终汇总报告（含失败原因）。"""
        success = sum(1 for _, ok, _ in results if ok)
        failed = sum(1 for _, ok, _ in results if not ok)
        total = success + failed

        lines = ["=" * 50]
        lines.append(f"  共 {total} 项  成功 {success}  失败 {failed}")
        lines.append("=" * 50)
        for desc, ok, err in results:
            status = "✅" if ok else "❌"
            line = f"  {status}  {desc}"
            if err:
                line += f"  [{err}]"
            lines.append(line)
        lines.append("=" * 50)
        lines.append(f"输出: {output_root}")

        report = "\n".join(lines)
        print(f"[ProductionExporter]\n{report}")

        if failed == 0:
            self._show_info("导出完成", report)
        elif success == 0:
            self._show_error("导出失败", report)
        else:
            self._show_warning(f"导出完成 ({success}/{total})", report)

    @staticmethod
    def _count_refs(refs: str) -> int:
        """从 KiCad 分组位号串推算元件数量。
        支持格式: "R1,R3,R5" / "C1,C3–C5,C7" / "U1"
        """
        if not refs or not refs.strip():
            return 0
        import re
        total = 0
        # 按逗号拆分
        parts = [p.strip() for p in refs.split(",")]
        for part in parts:
            if not part:
                continue
            # 匹配范围: 前缀+数字 – 前缀+数字，如 C4–C6 或 R3-R5
            m = re.match(r'^([A-Za-z]*)(\d+)\s*[–\-]\s*[A-Za-z]*(\d+)$', part)
            if m:
                lo, hi = int(m.group(2)), int(m.group(3))
                total += abs(hi - lo) + 1
            else:
                total += 1
        return total

    def _generate_bom_xlsx(self, csv_path: str, xlsx_path: str) -> tuple:
        """
        KiCad 内置 BOM 生成器已按 Value+Footprint 分组。
        本方法只需重整列为: PartNumber | Description | Quantity | Designator | Value | Footprint
        返回 (成功, 错误信息)。
        """
        try:
            import openpyxl
            import csv as csv_mod
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return False, "缺少 openpyxl 库，请在 KiCad Python 环境中安装: pip install openpyxl"

        try:
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                rows = list(csv_mod.reader(f))
            if len(rows) < 2:
                return False, "BOM CSV 为空"

            header = [h.strip().lower() for h in rows[0]]

            def col(*names):
                for n in names:
                    n = n.lower()
                    if n in header:
                        return header.index(n)
                return -1

            # 兼容中英文表头
            idx_ref = col("reference", "references", "designator", "designators",
                          "ref", "refs", "id", "位号")
            idx_val = col("value", "values", "值")
            idx_fp  = col("footprint", "footprints", "package", "封装")
            idx_qty = col("quantity", "qty", "count", "数量")
            idx_desc = col("description", "desc", "描述")
            idx_pn  = col("partnumber", "part_number", "mpn", "mfg_pn", "pn", "料号")
            idx_dnp = col("dnp")

            def get(row, idx):
                return row[idx].strip() if 0 <= idx < len(row) else ""

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM"
            ws.append(["PartNumber", "Description", "Quantity", "Designator", "Value", "Footprint", "DNP"])

            # 按 DNP 排序：非 DNP 在前，DNP 在后
            data_rows = rows[1:]
            if idx_dnp >= 0:
                data_rows.sort(key=lambda r: (get(r, idx_dnp) != "DNP", get(r, idx_ref)))

            for row in data_rows:
                if not row or all(c == "" for c in row):
                    continue
                refs = get(row, idx_ref)
                # Qty 列可能不存在、为空、或为 0（K8 不支持 QUANTITY 字段）
                raw_qty = get(row, idx_qty) if idx_qty >= 0 else ""
                qty = raw_qty if raw_qty and raw_qty != "0" else str(self._count_refs(refs))
                ws.append([
                    get(row, idx_pn),
                    get(row, idx_desc),
                    qty,
                    refs,
                    get(row, idx_val),
                    get(row, idx_fp),
                    get(row, idx_dnp),
                ])

            # 列宽 & 表头样式
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        length = sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value))
                        max_len = max(max_len, length)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

            hdr_font = Font(bold=True, size=11, color="FFFFFF")
            hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")

            wb.save(xlsx_path)
            print(f"[ProductionExporter] BOM XLSX: {xlsx_path}")
            return True, ""
        except Exception as e:
            msg = f"XLSX 转换失败: {e}"
            print(f"[ProductionExporter] {msg}")
            return False, msg


# ---- 注册插件 ----
try:
    ProductionExporter().register()
except Exception as e:
    # 重复注册时 KiCad 可能抛出异常，静默忽略
    pass
